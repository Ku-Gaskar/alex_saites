import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

START_HEADER_COLUMN_INDEX ='C' 
START_HEADER_ROW_INDEX = 3
START_DATA_ROW_INDEX = 9



class Exel_100():
    def __init__(self, filename):
        self.__wb: Workbook = None 
        try:
            self.__wb = openpyxl.load_workbook(filename)
        except IOError:
            print (f'Ошибка открітия файла: {filename}')
            return None
    
    def close(self):
        if  self.__wb:
            self.__wb.close()
    
    def readHeader(self,table_name:str) -> dict:
        ws:Worksheet=self.__wb.active
        list_columns={}
        for col in ws[START_HEADER_ROW_INDEX]:
            if col.column < column_index_from_string(START_HEADER_COLUMN_INDEX):
                continue
                 
            try:
                if col.value:
                    text=col.value.split('.')
                    if text[0] == table_name:
                        list_columns[col.column_letter]=text[1]
            except:
                print (f'Не правильный формат заголовка {col.value}' )
                return None
        return list_columns
    
    def readData(self,header:dict):
        ws:Worksheet=self.__wb.active
        head={key:vol[:vol.index(':')] for key,vol in header.items()}
        res=[]
        dim = ws.calculate_dimension().split(':')[1]
        for row in ws[START_HEADER_COLUMN_INDEX+str(START_DATA_ROW_INDEX): dim]:
            record={}
            for item in row:
                if item.value and (item.column_letter in head):
                    record[head[item.column_letter]]=item.value
            if record: 
                res.append(record)  
        return  res                


    